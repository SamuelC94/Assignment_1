from openpyxl import Workbook, load_workbook

workbook = load_workbook('saves/data.xlsx')
first_sheet = workbook.get_sheet_names()[0]
worksheet = workbook.get_sheet_by_name(first_sheet)
for row in worksheet.iter_rows():
    print(row)
    for cell in row:
        print(cell.value)

# check out the last row


